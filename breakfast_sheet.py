from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from pathlib import Path
import get_breakfast

def create_sheet(wb_name, sheet_name, b_data):
    WORKBOOK = load_workbook(filename=wb_name)
    WORKBOOK.create_sheet(sheet_name)
    SHEET = WORKBOOK[sheet_name]
    NUM_ROWS = len(b_data) + 1 # To include header row
    NUM_COLS = 13 # 7am - 7pm
    SHEET.insert_rows(idx = 1, amount = NUM_ROWS)
    SHEET.insert_cols(idx = 1, amount = NUM_COLS)
    # For column headers
    SHEET['B1'] = 'Total Sold'
    hour = 7 # Start at 7am
    # for col in SHEET.iter_cols(min_col=3, max_col=13):
    #     col[0].value = f'{hour}am' if hour < 12 else f'{hour - 1}pm'
    #     if hour == 12: # Exception
    #         col[0].value = '12pm'
    #     hour += 1
    # For row entries
    i = 0
    b_items = list(b_data.keys()) # Gets subscriptable breakfast item names
    for row in SHEET.iter_rows(min_row = 2, max_row = NUM_ROWS):
        row_name = row[0]
        b_item = b_items[i]
        i += 1
        row_name.value = b_item
        b_saletimes = b_data[b_item]
        row[1].value = len(b_saletimes) # Total number of the item sold
        for j in range(2, len(b_saletimes) + 2):
            row[j].value = b_saletimes[j - 2]
    SHEET.freeze_panes = f'C{NUM_ROWS + 1}'
    # To autoformat column widths and centering
    for col in SHEET.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal = 'center')
            cell.font = Font(name = 'Arial', size = 16)
            if len(str(cell.value)) > max_len:
                max_len = len(str(cell.value))
        adj_width = (max_len + 6) # Based on font size 16 for padding
        if col_letter == 'A': # Extra padding for breakfast item names because the for loop above doesn't work for it for some reason
            SHEET.column_dimensions[col_letter].width = adj_width + 5
        else:
            SHEET.column_dimensions[col_letter].width = adj_width
    SHEET[f'B1'].font = Font(name = 'Arial', size = 16, italic = True)
    for i in range(2, NUM_ROWS + 1):
        SHEET[f'B{i}'].font = Font(name = 'Arial', size = 16, bold = True)
    if 'Summary' in WORKBOOK.sheetnames:
        WORKBOOK.remove(WORKBOOK['Summary'])
    WORKBOOK.save(filename=wb_name)

def create_workbook(wb):
    workbook = Workbook()
    workbook.active.title = 'Summary'
    workbook.save(filename=wb)

def generate_daily_sheet(month, date):
    wb_name = f'{month}_Breakfast_Sale_Time.xlsx'
    b_path = Path(wb_name)
    if not b_path.exists():
        create_workbook(wb_name)
    b_date = f'{date[4:6]}_{date[6:]}_{date[2:4]}'
    b_data = get_breakfast.get_check_data(f'{date}0700', f'{date}1900')
    # b_data #
    # Key: Breakfast item names
    # Values: Sale times
    get_breakfast.make_breakfast_text_file(b_date, b_data)
    create_sheet(wb_name, f'{date[4:6]}_{date[6:]}_{date[:4]}', b_data)

generate_daily_sheet('January', '20260125')

