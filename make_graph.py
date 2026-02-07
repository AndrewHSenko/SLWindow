from openpyxl import load_workbook
from openpyxl.chart import (LineChart, Reference)
from openpyxl.chart.axis import ChartLines

def make_daily_prod(wbook, prods, sheet_name, graph_name, legend_name, start_hour, ylimit=40, smooth_it=False):
    WORKBOOK = load_workbook(filename=wbook)
#    WS = WORKBOOK[time.strftime('%m_%d_%Y')]
    day = sheet_name + ' Graph'
    WORKBOOK.create_sheet(day)
    WS = WORKBOOK[day]
    fives_index = 0
    fives_data = [int(float(prods[prod])) for prod in prods]
    for _ in range(10):
        fives_data.append(0)
    WS.insert_cols(idx = 1)
    hr_start = hr_end = start_hour
    min_start = 0
    min_end = 5 if not smooth_it else 10 # intervals of 10 mins for smoothed
    WS.insert_rows(idx = 1, amount = len(fives_data) + 10)
    for row in WS.iter_rows(min_row = 2):
        if (hr_start == 20 and min_start == 0) or fives_index == len(fives_data):
            break
        header = row[0]
        if min_end == 60:
            min_end = 0
            hr_end += 1
        if min_start == 60:
            min_start = 0
            hr_start += 1
        header.value = '{hour_start}:{minute_start}{tod_start}-{hour_end}:{minute_end}{tod_end}'.format(
            hour_start = (hr_start - 12) if hr_start > 12 else hr_start,
            minute_start = f'0{min_start}' if min_start < 10 else min_start,
            tod_start = 'AM' if hr_start < 12 else 'PM',
            hour_end = f'{hr_end - 12}' if hr_end > 12 else hr_end,
            minute_end = f'0{min_end}' if min_end < 10 else min_end,
            tod_end = 'AM' if hr_end < 12 else 'PM'
        )
        min_start += 5
        min_end += 5
        row[1].value = fives_data[fives_index]
        fives_index += 1
    WS['B1'] = legend_name
    c1 = LineChart()
    c1.title = graph_name + ' 5 min Production'
    c1.style = 13
    c1.height = 15
    c1.width = 30
    data = Reference(WS, min_row = 1, min_col = 2, max_row = len(fives_data), max_col = 2)
    cats = Reference(WS, min_row = 2, min_col = 1, max_row = len(fives_data), max_col = 1)
    c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats)
    c1.x_axis.tickLblSkip = 6
    c1.x_axis.tickMarkSkip = 6
    c1.x_axis.majorGridlines = ChartLines()
    c1.x_axis.minorGridlines = ChartLines()
    c1.y_axis.scaling.min = 0
    c1.y_axis.scaling.max = ylimit
    line = c1.series[0]
    line.smooth = True
    WS.add_chart(c1, 'D1')
    WORKBOOK.save(filename=wbook)