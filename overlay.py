from openpyxl import Workbook, load_workbook
from openpyxl.chart import (LineChart, BarChart, Reference)
from openpyxl.chart.axis import ChartLines
import get_pu_window as pu
from os import mkdir

def ratio(sl_prods, foh_prods): # assumes sl_prods and foh_prods have the same intervals
    ratios = {}
    running_count = 0
    i = 0
    for sh, sp in sl_prods.items():
        if i == 0:
            ratios[i] = int(foh_prods[sh]) - int(float(sp))
            running_count = int(foh_prods[sh]) - int(float(sp))
        else:
            running_count = running_count + (int(foh_prods[sh]) - int(float(sp)))
            ratios[i] = running_count
        i += 1
    return ratios

def create_overlay(wbook, sl_prods, foh_prods, foh_window, foh_actual, sheet_name, graph_name, first_data_name, sec_data_name, start_hour, ylimit=40):
    try:
        WORKBOOK = load_workbook(filename=wbook)
    #    WS = WORKBOOK[time.strftime('%m_%d_%Y')]
        day = sheet_name + ' Overlay Graph'
        WORKBOOK.create_sheet(day)
        WS = WORKBOOK[day]
        fives_index = 0
        if sl_prods:
            LENGTH = len(sl_prods)
        elif foh_prods:
            LENGTH = len(foh_prods)
        else:
            LENGTH = len(foh_window)
        if sl_prods:
            sl_data = [int(float(sl_prods[prod])) for prod in sl_prods]
            for i in range(10):
                sl_data.append(0)
        if foh_prods:
            foh_data = [int(float(foh_prods[prod])) for prod in foh_prods]
            for i in range(10):
                foh_data.append(0)
        if foh_window:
            foh_w = [int(float(foh_window[prod])) for prod in foh_window]
            for i in range(10):
                foh_w.append(0)
        if foh_actual:
            foh_a = [int(float(foh_actual[prod])) for prod in foh_actual]
            for i in range(10):
                foh_a.append(0)
        WS.insert_cols(idx = 1, amount=4)
        hr_start = hr_end = start_hour
        min_start = 0
        min_end = 5
        WS.insert_rows(idx = 1, amount = LENGTH + 10)
        for row in WS.iter_rows(min_row = 2):
            if (hr_start == 20 and min_start == 0) or fives_index == LENGTH:
                break
            header = row[0]
            if min_end == 60:
                min_end = 0
                hr_end += 1
            if min_start == 60:
                min_start = 0
                hr_start += 1
            header.value = '{hour_start}:{minute_start}{tod_start}-{hour_end}:{minute_end}{tod_end}'.format(
                hour_start = (hr_start - 12) if hr_start > 12 else hr_start,
                minute_start = f'0{min_start}' if min_start < 10 else min_start,
                tod_start = 'AM' if hr_start < 12 else 'PM',
                hour_end = f'{hr_end - 12}' if hr_end > 12 else hr_end,
                minute_end = f'0{min_end}' if min_end < 10 else min_end,
                tod_end = 'AM' if hr_end < 12 else 'PM'
            )
            min_start += 5
            min_end += 5
            if sl_prods:
                row[1].value = sl_data[fives_index]
            if foh_prods:
                row[2].value = foh_data[fives_index]
            if foh_window:
                row[3].value = foh_w[fives_index]
            if foh_actual:
                row[4].value = foh_a[fives_index]
            fives_index += 1
        WS['B1'] = first_data_name
        WS['C1'] = sec_data_name
        WS['D1'] = 'FoH Planned (Window)'
        WS['E1'] = 'FoH Actual (Window)'
        # For the SL production from the KDS and FoH entries in Squirrel
        c1 = LineChart()
        c1.title = graph_name + ' 5 Min Window'
        c1.style = 13
        c1.height = 15
        c1.width = 30
        data = Reference(WS, min_row = 1, min_col = 2, max_row = LENGTH, max_col = 3) # Removed 5 for bar chart
        cats = Reference(WS, min_row = 2, min_col = 1, max_row = LENGTH, max_col = 1)
        c1.add_data(data, titles_from_data=True)
        c1.set_categories(cats)
        c1.x_axis.tickLblSkip = 6
        c1.x_axis.tickMarkSkip = 6
        c1.x_axis.majorGridlines = ChartLines()
        c1.x_axis.minorGridlines = ChartLines()
        c1.y_axis.scaling.min = 0
        c1.y_axis.scaling.max = ylimit
        line = c1.series[0]
        line.smooth = True
        # For the FoH Window and FoH Actual from Google Sheets
        c2 = BarChart()
        w_data = Reference(WS, min_row = 1, min_col = 4, max_row = LENGTH)
        c2.add_data(w_data, titles_from_data=True)
        c3 = BarChart()
        a_data = Reference(WS, min_row = 1, min_col = 5, max_row = LENGTH)
        c3.add_data(a_data, titles_from_data=True)
        c1 += c2
        c1 += c3
        WS.add_chart(c1, 'F1')
        WORKBOOK.save(filename=wbook)
    except Exception as e:
        with open('overlay_errors.txt', 'a') as err_file:
            err_file.write(f'{e}\n')