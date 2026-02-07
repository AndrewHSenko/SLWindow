from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# Maybe add column for best production per hour?
def create_sheet(workbook_name, s_name, data, num_rows, start_hour):
    WORKBOOK = load_workbook(filename=workbook_name)
    WORKBOOK.create_sheet(s_name)
    SHEET = WORKBOOK[s_name]
    NUM_ROWS = num_rows
    NUM_COLS = 14
    SHEET.insert_rows(idx = 1, amount = NUM_ROWS)
    SHEET.insert_cols(idx = 1, amount = NUM_COLS)
    # For column headers
    SHEET['B1'] = 'Total'
    minute = 0
    end_min = minute + 5
    for col in SHEET.iter_cols(min_col = 3, max_col = 14):
        col[0].value = 'xx:{minute}-xx:{end_min}'.format(minute = f'0{minute}' if minute < 10 else minute, end_min = f'0{end_min}' if end_min < 10 else end_min) # Col header cells
        minute += 5
        end_min = minute + 5
    # For row filling
    hour = start_hour
    PRODS = data
    prods_i = 0
    for row in SHEET.iter_rows(min_row = 2, max_row = NUM_ROWS - 1):
        # Handles row headers (Why use another loop outside of this one to do row headers when we already grab every data row here?)
        row_header = row[0]
        if hour >= start_hour and hour < 12:
            row_header.value = f'{hour}AM'
            hour += 1
        elif hour == 12:
            row_header.value = '12PM'
            hour = 1
        else:
            row_header.value = f'{hour}PM'
            hour += 1
        # Fills "total" column
        total_col_cell = row[1]
        total_col_cell.value = f'=SUM(C{row[0].row}:N{row[0].row})'
        # Fills data for each row
        for i in range(2, NUM_COLS): # From start of data column to last data column
            if prods_i >= len(PRODS): # No more data to use to fill
                row[i].value = 0
            else:
                row[i].value = PRODS[prods_i]
                prods_i += 1
    total_cell_name = f'A{NUM_ROWS}'
    total_cell_data = f'B{NUM_ROWS}'
    SHEET[total_cell_name] = 'Total'
    SHEET[total_cell_data] = f'=SUM(B2:B{NUM_ROWS - 1})' # Last row of data
    SHEET.freeze_panes = f'C{NUM_ROWS}'
    # To autoformat column widths and centering
    for col in SHEET.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(horizontal = 'center')
            cell.font = Font(name = 'Arial', size = 16)
            if len(str(cell.value)) > max_len:
                max_len = len(str(cell.value))
            else:
                continue
        adj_width = (max_len + 6) # Based on font size 16
        SHEET.column_dimensions[col_letter].width = adj_width
    SHEET[total_cell_data].font = Font(name = 'Arial', size = 16, bold = True)
    if 'Button_Data' not in workbook_name: # daily output file
        if 'Summary' in WORKBOOK.sheetnames:
            WORKBOOK.remove(WORKBOOK['Summary'])
    WORKBOOK.save(filename=workbook_name)

def create_workbook(wb):
    workbook = Workbook()
    workbook.active.title = 'Summary'
    workbook.save(filename=wb)

def generate_daily_sheet(wb_name, data, new_wb, sheet_name, num_rows, start_hour):
    if new_wb:
        create_workbook(wb_name)
    just_prods = [int(float(data[d])) for d in data]
    create_sheet(wb_name, sheet_name, just_prods, num_rows, start_hour)